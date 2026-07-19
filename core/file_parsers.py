"""
File parsers — read trial balance from Excel, CSV, or PDF.

Returns a list of dicts: {code, name, debit, credit}

Strategy:
  - Excel/CSV: best support, with auto header detection
  - PDF: three-tier strategy:
      1. extract_tables() with balance-column detection
      2. line-by-line text extraction (with balance preference)
      3. OCR fallback for scanned/image PDFs (Arabic supported)
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Optional

from .arabic_utils import clean, to_western_digits, has_arabic


def parse_file(path: str) -> list[dict]:
    p = Path(path)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return parse_excel(path)
    if ext == ".pdf":
        return parse_pdf(path)
    if ext == ".csv":
        return parse_csv(path)
    raise ValueError(f"نوع الملف غير مدعوم: {ext}")


def _is_summary_line(name: str) -> bool:
    if not name:
        return True
    n = name.strip()
    low = n.lower()
    if low.startswith("total ") or low == "total" or low.startswith("subtotal"):
        return True
    if "total" in low and len(n) < 30:
        return True
    if low.startswith("page:") or "printed by" in low or "printed on" in low:
        return True
    if n.startswith("إجمالي") or n.startswith("المجموع") or n.startswith("الرصيد"):
        return True
    if n.startswith("Total ") or n.startswith("Total:"):
        return True
    if n.startswith("#"):
        return True
    return False


def _fix_arabic_spacing(text: str) -> str:
    if not text:
        return text
    # Remove kashida (tatweel) character ـ
    text = text.replace('ـ', '')
    text = text.replace('ـــ', '')
    # Remove spaces between Arabic letters
    text = re.sub(r'([؀-ۿ])\s+([؀-ۿ])', r'\1\2', text)
    # Clean up multiple spaces
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _is_number(s) -> bool:
    if s is None:
        return False
    try:
        float(str(s).replace(",", "").replace(" ", ""))
        return True
    except (ValueError, TypeError):
        return False


def _num(s) -> float:
    if s is None or s == "":
        return 0.0
    try:
        s = str(s).replace(",", "").replace(" ", "").replace("٬", "").replace("٫", ".")
        s = to_western_digits(s)
        return float(s)
    except (ValueError, TypeError):
        return 0.0


# ──────────────────────────────────────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────────────────────────────────────

def parse_excel(path: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError("لم يتم العثور على صف رأس الجدول في الملف")

    header = [clean(c) for c in rows[header_idx]]
    col_map = _map_columns(header)
    if "name" not in col_map:
        raise ValueError("الملف لا يحتوي على عمود اسم الحساب")

    out: list[dict] = []
    for r in rows[header_idx + 1:]:
        if not r or all(c is None or clean(c) == "" for c in r):
            continue
        code = clean(r[col_map["code"]]) if "code" in col_map and col_map["code"] < len(r) else ""
        name = clean(r[col_map["name"]]) if "name" in col_map and col_map["name"] < len(r) else ""
        if not name or _is_summary_line(name):
            continue
        balance = 0
        if "balance" in col_map and col_map["balance"] < len(r):
            balance = _num(r[col_map["balance"]])
        elif "debit" in col_map and "credit" in col_map:
            d = _num(r[col_map["debit"]]) if col_map["debit"] < len(r) else 0
            c = _num(r[col_map["credit"]]) if col_map["credit"] < len(r) else 0
            balance = d - c
        out.append({"code": code, "name": name, "debit": max(balance, 0), "credit": max(-balance, 0)})
    return out


def _find_header_row(rows):
    keys = ("الحساب", "اسم", "رمز", "مدين", "دائن", "رصيد", "افتتاحي",
            "Account", "Debit", "Credit", "Code", "Name", "Balance")
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        joined = " ".join(clean(c) for c in row if c is not None)
        if sum(1 for k in keys if k in joined) >= 3:
            return i
    return None


def _map_columns(header):
    mapping = {}
    for i, h in enumerate(header):
        if not h:
            continue
        h_low = h.lower()
        if "code" in h_low or "رقم" in h or "رمز" in h or "كود" in h:
            mapping.setdefault("code", i)
        elif "account" in h_low or "اسم" in h or "حساب" in h or "البيان" in h:
            mapping.setdefault("name", i)
        elif "balance" in h_low or "رصيد" in h or "الرصيد" in h:
            mapping.setdefault("balance", i)
        elif "debit" in h_low or "مدين" in h:
            mapping.setdefault("debit", i)
        elif "credit" in h_low or "دائن" in h:
            mapping.setdefault("credit", i)
    return mapping


# ──────────────────────────────────────────────────────────────────────────────
# PDF
# ──────────────────────────────────────────────────────────────────────────────

def parse_pdf(path: str) -> list[dict]:
    out = []
    try:
        out = _parse_pdf_tables(path)
        if out and len(out) >= 3:
            return _finalize(out)
    except Exception:
        pass
    try:
        out = _parse_pdf_text(path)
        if out and len(out) >= 3:
            return _finalize(out)
    except Exception:
        pass
    try:
        out = _parse_pdf_ocr(path)
        if out and len(out) >= 3:
            return _finalize(out)
    except Exception:
        pass
    if not out:
        raise ValueError(
            "تعذر قراءة الملف. الأسباب المحتملة:\n"
            "1. الملف محمي بكلمة مرور\n"
            "2. الملف ليس ميزان مراجعة\n"
            "3. صيغة غير مدعومة\n"
            "الحل: جرّب رفع ملف Excel (.xlsx) أو CSV بدلاً من PDF"
        )
    return _finalize(out)


def _parse_pdf_tables(path: str) -> list[dict]:
    import pdfplumber
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            try:
                tables = page.extract_tables() or []
                for tbl in tables:
                    if not tbl or len(tbl) < 3:
                        continue
                    header_idx = _find_pdf_header(tbl)
                    if header_idx is None:
                        continue
                    col_map = _map_pdf_columns(tbl[header_idx])
                    if "name" not in col_map and "balance" not in col_map and "debit" not in col_map:
                        continue
                    for r in tbl[header_idx + 1:]:
                        row = _row_from_table_smart(r, col_map)
                        if row:
                            out.append(row)
            except Exception:
                continue
    return out


def _parse_pdf_text(path: str) -> list[dict]:
    import pdfplumber
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                row = _parse_pdf_line_smart(line)
                if row:
                    out.append(row)
    return out


def _parse_pdf_ocr(path: str) -> list[dict]:
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except ImportError:
        return []
    out = []
    try:
        page_count = 5
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                page_count = min(len(pdf.pages), 10)
        except Exception:
            pass
        images = convert_from_path(path, dpi=150, first_page=1, last_page=page_count)
        for img in images:
            try:
                text = pytesseract.image_to_string(img, lang='ara+eng', config='--psm 6')
            except Exception:
                text = pytesseract.image_to_string(img, lang='eng', config='--psm 6')
            for line in text.splitlines():
                row = _parse_pdf_line_smart(line)
                if row:
                    out.append(row)
    except Exception:
        return []
    return out


def _find_pdf_header(tbl):
    keys = ("الحساب", "اسم", "رمز", "مدين", "دائن", "رصيد", "Account", "Debit", "Credit", "Code", "Name", "Balance", "OB")
    for i, row in enumerate(tbl[:5]):
        joined = " ".join((c or "") for c in row if c)
        if sum(1 for k in keys if k in joined) >= 2:
            return i
    return None


def _map_pdf_columns(header):
    mapping = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        h_low = str(h).lower()
        h_str = str(h)
        if "code" in h_low or "رمز" in h_str or "كود" in h_str or "رقم" in h_str:
            mapping.setdefault("code", i)
        elif "account" in h_low or "اسم" in h_str or "حساب" in h_str:
            mapping.setdefault("name", i)
        elif "balance" in h_low or "رصيد" in h_str or "الرصيد" in h_str:
            mapping.setdefault("balance", i)
        elif "ob" == h_low or "افتتاحي" in h_str:
            mapping.setdefault("ob", i)
        elif "debit" in h_low or "مدين" in h_str:
            mapping.setdefault("debit", i)
        elif "credit" in h_low or "دائن" in h_str:
            mapping.setdefault("credit", i)
    return mapping


def _row_from_table_smart(r, col_map):
    if not r or all((c is None or clean(str(c)) == "") for c in r):
        return None

    # Filter Total/summary lines BEFORE extracting data
    raw_text = " ".join(str(c) for c in r if c)
    if re.match(r"^[\s]*(Total|إجمالي|المجموع|Subtotal)\s+", raw_text, re.IGNORECASE):
        return None
    if r and r[0] and (str(r[0]).lower().startswith("total") or "إجمالي" in str(r[0])):
        return None

    # Get name
    if "name" in col_map and col_map["name"] < len(r) and r[col_map["name"]]:
        name = clean(r[col_map["name"]])
    else:
        name = ""
        for c in r:
            if c and not _is_number(c) and clean(str(c)):
                name = clean(str(c))
                break

    if not name or _is_summary_line(name):
        return None

    code = ""
    if "code" in col_map and col_map["code"] < len(r) and r[col_map["code"]]:
        code = clean(r[col_map["code"]])

    # Extract balance - prefer balance column, then last numeric
    balance = 0.0
    if "balance" in col_map and col_map["balance"] < len(r) and r[col_map["balance"]]:
        balance = _num(r[col_map["balance"]])
    elif "debit" in col_map and "credit" in col_map:
        d = _num(r[col_map["debit"]]) if col_map["debit"] < len(r) and r[col_map["debit"]] else 0
        c = _num(r[col_map["credit"]]) if col_map["credit"] < len(r) and r[col_map["credit"]] else 0
        if d != 0 and c == 0:
            balance = d
        elif c != 0 and d == 0:
            balance = -c
        elif d != 0 and c != 0:
            balance = d - c
    elif "debit" in col_map and col_map["debit"] < len(r) and r[col_map["debit"]]:
        balance = _num(r[col_map["debit"]])
    elif "credit" in col_map and col_map["credit"] < len(r) and r[col_map["credit"]]:
        balance = -_num(r[col_map["credit"]])

    if abs(balance) < 0.001:
        return None

    return {
        "code": code,
        "name": name,
        "debit": balance if balance > 0 else 0,
        "credit": -balance if balance < 0 else 0,
    }


def _parse_pdf_line_smart(line: str):
    line = clean(line)
    if not line or len(line) < 4:
        return None
    if _is_summary_line(line):
        return None
    # Filter Total/Subtotal/header lines BEFORE extraction
    if re.match(r"^[\s]*(Total|إجمالي|المجموع|Subtotal)\s+", line, re.IGNORECASE):
        return None
    # Match "Total 12,345 67,890" (just totals)
    if re.match(r"^[\s]*(Total|إجمالي)\s+[\d,.]+\s+[\d,.]+\s*$", line, re.IGNORECASE):
        return None

    # Extract numbers
    western = to_western_digits(line)
    numbers = re.findall(r'-?[\d,]+\.?\d*', western)
    valid_numbers = []
    for n in numbers:
        try:
            v = float(n.replace(",", ""))
            if abs(v) > 0.001:
                valid_numbers.append(v)
        except ValueError:
            continue

    if not valid_numbers:
        return None

    # Extract code (first 4+ digit number at start of line)
    code = ""
    code_match = re.match(r'^\s*(\d{4,})\b', western)
    if code_match:
        code = code_match.group(1)

    # Remove the code from valid_numbers to avoid using it as balance
    if code:
        try:
            code_as_num = float(code)
            if valid_numbers and abs(valid_numbers[0] - code_as_num) < 0.01:
                valid_numbers = valid_numbers[1:]
        except ValueError:
            pass

    if not valid_numbers:
        return None

    # Also detect short codes (section headers in SAP: 101, 102, 201, etc.)
    # These are 1-6 digits and represent account categories, not actual accounts
    if not code:
        short_code = re.match(r'^\s*(\d{1,6})\s*[-–]', western)
        if short_code:
            return None

    # Filter out section headers (codes < 7 digits typically section headers in SAP)
    # e.g. "101 - إجمالي المتداولة" — code 101 is a section, not an account
    if code and len(code) < 7:
        # This is likely a section header. Skip it.
        return None

    # Extract name (remove leading code, then remove all numbers)
    name = line
    name = re.sub(r'^\s*\d+\s*[-–]?\s*', '', name)
    name_no_nums = re.sub(r'-?[\d,]+\.?\d*', '', western)
    name_no_nums = clean(name_no_nums)
    if not name_no_nums or len(name_no_nums) < 2:
        return None
    name = name_no_nums

    # Skip page headers/footers
    skip_patterns = ("Posting Date", "Template:", "Cycle:", "Printed By", "Printed On",
                     "Page:", "BP:", "Customer Group", "Supplier Group", "Local Currency",
                     "Currency", "From", "To", "Trial Balance", "Annual Report")
    for p in skip_patterns:
        if p in name:
            return None

    # Determine balance - use LAST number (it's usually the closing balance)
    if len(valid_numbers) >= 1:
        balance = valid_numbers[-1]
    else:
        balance = 0

    if abs(balance) < 0.001:
        return None

    return {
        "code": code,
        "name": name,
        "debit": balance if balance > 0 else 0,
        "credit": -balance if balance < 0 else 0,
    }


def _finalize(rows):
    seen = set()
    out = []
    for r in rows:
        name = _fix_arabic_spacing(r["name"])
        if _is_summary_line(name) or len(name) < 2:
            continue
        key = name.strip()
        if key in seen:
            continue
        seen.add(key)
        code = r.get("code", "")
        if not code:
            m = re.search(r'(\d{4,})', name)
            if m:
                code = m.group(1)
        out.append({
            "code": code,
            "name": name,
            "debit": abs(r.get("debit", 0)),
            "credit": abs(r.get("credit", 0)),
        })
    return out


# ──────────────────────────────────────────────────────────────────────────────
# CSV
# ──────────────────────────────────────────────────────────────────────────────

def parse_csv(path: str) -> list[dict]:
    rows = []
    with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append([clean(c) for c in row])
    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError("لم يتم العثور على صف الرأس")
    header = rows[header_idx]
    col_map = _map_columns(header)
    if "name" not in col_map:
        raise ValueError("الملف لا يحتوي على عمود اسم الحساب")
    out = []
    for r in rows[header_idx + 1:]:
        if not r or all(c == "" for c in r):
            continue
        code = r[col_map["code"]] if "code" in col_map and col_map["code"] < len(r) else ""
        name = r[col_map["name"]] if "name" in col_map and col_map["name"] < len(r) else ""
        if not name or _is_summary_line(name):
            continue
        balance = 0
        if "balance" in col_map and col_map["balance"] < len(r):
            balance = _num(r[col_map["balance"]])
        elif "debit" in col_map and "credit" in col_map:
            d = _num(r[col_map["debit"]]) if col_map["debit"] < len(r) else 0
            c = _num(r[col_map["credit"]]) if col_map["credit"] < len(r) else 0
            balance = d - c
        out.append({"code": code, "name": name, "debit": max(balance, 0), "credit": max(-balance, 0)})
    return out
