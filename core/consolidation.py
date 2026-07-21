"""
Consolidated Financial Statements Engine (IFRS 10).

Supports:
  1. Full consolidation of subsidiaries (100% line-by-line).
  2. Non-Controlling Interest (NCI) computation.
  3. Intercompany (IC) eliminations: balances, revenues/expenses, and
     investment-in-subsidiary vs. equity elimination.

Each input is a parsed trial balance job. The output is a unified set of
financial statements ready to render or export.

Schema of the group definition (`group`):
  {
    "id": "grp_abc",
    "name": "مجموعة بهاء الدين",
    "parent_company_id": "c_xyz",        # الشركة الأم
    "created_at": 1700000000,
  }

Each ownership link:
  {
    "group_id": "grp_abc",
    "company_id": "c_xyz",
    "parent_company_id": "c_xyz",        # لمن تعود ملكيتها
    "ownership_pct": 100.0,              # 0..100
    "consolidation_method": "full",      # full | proportional | equity
  }
"""

from __future__ import annotations

from typing import Optional


# ──────────────────────────────────────────────────────────────────────────────
# 1) Line-level aggregation
# ──────────────────────────────────────────────────────────────────────────────

def _label_map(stmt: dict) -> dict[str, dict]:
    """Index a statement's lines by their label for fast lookup."""
    out = {}
    for line in (stmt.get("lines") or []):
        out[line.get("label", "")] = line
    return out


def _get_company_amount(jobs_data: list[dict], stmt_key: str, label: str) -> float:
    """Sum the same label across all companies in the group."""
    total = 0.0
    for jd in jobs_data:
        stmt = (jd.get("statements") or {}).get(stmt_key) or {}
        for line in (stmt.get("lines") or []):
            if line.get("label") == label:
                total += float(line.get("amount", 0) or 0)
                break
    return total


def _aggregate_lines(jobs_data: list[dict], stmt_key: str, ownership_links: list[dict]) -> list[dict]:
    """
    Aggregate one statement (e.g. balance_sheet) across companies.

    Logic per IFRS 10 full consolidation:
      - Every subsidiary's full amount (100%) is brought in.
      - NCI is calculated separately on net assets/equity, NOT by scaling
        individual lines here.
      - The parent company itself is included at 100% (no scaling).

    Returns the aggregated lines (label, amount, indent, bold) ready to render.
    """
    if not jobs_data:
        return []
    # use parent company statement as the master template (labels + order)
    master_stmt = (jobs_data[0].get("statements") or {}).get(stmt_key) or {}
    master_lines = master_stmt.get("lines") or []

    aggregated = []
    for line in master_lines:
        label = line.get("label", "")
        bold = line.get("bold", False)
        indent = line.get("indent", 0)
        total = 0.0
        for jd in jobs_data:
            stmt = (jd.get("statements") or {}).get(stmt_key) or {}
            for ln in (stmt.get("lines") or []):
                if ln.get("label") == label:
                    total += float(ln.get("amount", 0) or 0)
                    break
        aggregated.append({
            "label": label,
            "amount": round(total, 2),
            "bold": bold,
            "indent": indent,
        })
    return aggregated


# ──────────────────────────────────────────────────────────────────────────────
# 2) Non-Controlling Interest (NCI)
# ──────────────────────────────────────────────────────────────────────────────

def compute_nci(equity_aggregated: float, ownership_pct: float) -> dict:
    """
    NCI = aggregated equity × (1 - ownership_pct/100)
    Parent share = aggregated equity × (ownership_pct/100)

    For full consolidation of a wholly-owned sub, NCI = 0.
    For a 60% owned sub: parent gets 60% of net assets, NCI gets 40%.
    """
    pct = max(0.0, min(100.0, ownership_pct or 0.0))
    parent_share = round(equity_aggregated * (pct / 100.0), 2)
    nci = round(equity_aggregated - parent_share, 2)
    return {
        "ownership_pct": pct,
        "parent_share": parent_share,
        "nci": nci,
    }


# ──────────────────────────────────────────────────────────────────────────────
# 3) Intercompany detection
# ──────────────────────────────────────────────────────────────────────────────

# Sub-categories that the trial balance marks as intercompany
IC_CATEGORIES = {
    "ic_receivable",          # مدينون بين شركات المجموعة
    "ic_payable",             # دائنون بين شركات المجموعة
    "ic_revenue",             # إيرادات متبادلة
    "ic_expense",             # مصاريف متبادلة
    "ic_loan_receivable",     # قروض ممنوحة لتابعة
    "ic_loan_payable",        # قروض مستلمة من تابعة
    "investment_in_sub",      # استثمار الشركة الأم في التابعة
    "ic_dividend_receivable", # توزيعات مستحقة من تابعة
    "ic_dividend_payable",    # توزيعات مستحقة لشركة أم
}


def detect_intercompany(jobs_data: list[dict]) -> dict:
    """
    Walk every company's accounts and collect those flagged as intercompany.
    Returns: {
      "ic_receivable": total_amt,        # مدينون متبادلون
      "ic_payable":    -total_amt,       # دائنون متبادلون
      "ic_revenue":    total_amt,
      "ic_expense":    -total_amt,
      "ic_loan_receivable": total_amt,
      "ic_loan_payable":    -total_amt,
      "investment_in_sub":  total_amt,   # في ميزان الأم
      "subs_equity_elim":   -total_amt,  # حقوق ملكية التابعة (تُحذف من المجموعة)
    }
    """
    totals = {
        "ic_receivable": 0.0,
        "ic_payable": 0.0,
        "ic_revenue": 0.0,
        "ic_expense": 0.0,
        "ic_loan_receivable": 0.0,
        "ic_loan_payable": 0.0,
        "investment_in_sub": 0.0,
        "ic_dividend_receivable": 0.0,
        "ic_dividend_payable": 0.0,
    }
    for jd in jobs_data:
        for a in (jd.get("accounts") or jd.get("raw_rows") or []):
            cat = a.get("sub_category") or a.get("category") or ""
            amt = float(a.get("balance", 0) or a.get("amount", 0) or 0)
            if cat in totals:
                totals[cat] = round(totals[cat] + amt, 2)
            elif cat in ("intercompany_receivable",):
                totals["ic_receivable"] = round(totals["ic_receivable"] + amt, 2)
            elif cat in ("intercompany_payable",):
                totals["ic_payable"] = round(totals["ic_payable"] + amt, 2)
    return totals


# ──────────────────────────────────────────────────────────────────────────────
# 4) Eliminations
# ──────────────────────────────────────────────────────────────────────────────

def apply_eliminations(consolidated: dict, ic_totals: dict) -> dict:
    """
    Apply intercompany eliminations to the aggregated statements.

    IFRS 10 required eliminations:
      - Receivables vs. payables (دائنون متبادلون)
      - Revenue vs. expense (إيرادات/مصاريف متبادلة)
      - Investment in subsidiary vs. equity at acquisition
        + NCI recognition (handled separately via compute_nci)
      - Intra-group dividends payable/receivable

    Each line whose label matches the pattern below is reduced by the
    corresponding IC total.
    """
    KEYWORD_MAP = {
        "المدينون":        "ic_receivable",
        "مدينون":          "ic_receivable",
        "ذمم مدينة":       "ic_receivable",
        "الدائنون":        "ic_payable",
        "دائنون":          "ic_payable",
        "ذمم دائنة":       "ic_payable",
        "إيرادات":         "ic_revenue",
        "المبيعات":        "ic_revenue",
        "ايرادات":         "ic_revenue",
        "مصاريف":          "ic_expense",
        "مصروفات":         "ic_expense",
        "تكلفة":           "ic_expense",
        "استثمارات":       "investment_in_sub",
        "الاستثمارات":     "investment_in_sub",
        "استثمار في":      "investment_in_sub",
    }

    def _eliminate_in_lines(lines: list[dict]) -> list[dict]:
        new = []
        for line in lines:
            label = (line.get("label") or "").strip()
            amount = float(line.get("amount", 0) or 0)
            applied = False
            for kw, ic_key in KEYWORD_MAP.items():
                if kw in label and ic_totals.get(ic_key, 0) != 0:
                    amount = round(amount - ic_totals[ic_key], 2)
                    applied = True
                    break
            new_line = dict(line)
            new_line["amount"] = amount
            new_line["eliminated"] = applied
            new.append(new_line)
        return new

    for stmt_key, stmt in consolidated.get("statements", {}).items():
        if isinstance(stmt, dict) and "lines" in stmt:
            stmt["lines"] = _eliminate_in_lines(stmt["lines"])
    consolidated["ic_eliminations"] = ic_totals
    return consolidated


# ──────────────────────────────────────────────────────────────────────────────
# 5) Full consolidation pipeline
# ──────────────────────────────────────────────────────────────────────────────

def consolidate(
    group: dict,
    ownership_links: list[dict],
    jobs_data: list[dict],
) -> dict:
    """
    Main pipeline: aggregate → eliminate → compute NCI.
    jobs_data: list of parsed-job dicts (one per company in the group)
               Each must include 'company_id', 'company_name', 'period',
               and the parsed 'statements' dict.
    """
    parent_id = group.get("parent_company_id")
    companies = []
    for jd in jobs_data:
        companies.append({
            "company_id": jd.get("company_id"),
            "company_name": jd.get("company_name") or jd.get("company", ""),
            "period": jd.get("period", ""),
        })

    # Average ownership of subsidiaries (for NCI display)
    sub_links = [l for l in ownership_links if l.get("company_id") != parent_id]
    avg_ownership = 0.0
    if sub_links:
        avg_ownership = sum(l.get("ownership_pct", 0) for l in sub_links) / len(sub_links)

    # 1. Aggregate statements
    aggregated_statements = {}
    stmt_keys = set()
    for jd in jobs_data:
        stmt_keys.update((jd.get("statements") or {}).keys())

    for stmt_key in stmt_keys:
        aggregated_statements[stmt_key] = {
            "title": _first_title(jobs_data, stmt_key) or stmt_key,
            "lines": _aggregate_lines(jobs_data, stmt_key, ownership_links),
        }

    # 2. Detect & apply intercompany eliminations
    ic_totals = detect_intercompany(jobs_data)
    consolidated = {
        "group": group,
        "companies": companies,
        "ownership_links": ownership_links,
        "avg_ownership_pct": round(avg_ownership, 2),
        "statements": aggregated_statements,
    }
    consolidated = apply_eliminations(consolidated, ic_totals)

    # 3. NCI: based on equity
    total_equity = _get_company_amount(jobs_data, "balance_sheet", "حقوق الملكية")
    # Adjust by eliminations already applied
    for line in (consolidated["statements"].get("balance_sheet", {}).get("lines") or []):
        if "حقوق الملكية" in (line.get("label") or "") and line.get("bold"):
            total_equity = float(line.get("amount", 0) or 0)
            break

    nci_info = compute_nci(total_equity, avg_ownership)
    consolidated["nci"] = nci_info
    consolidated["nci"]["equity_total"] = total_equity

    return consolidated


def _first_title(jobs_data: list[dict], stmt_key: str) -> str:
    for jd in jobs_data:
        stmt = (jd.get("statements") or {}).get(stmt_key) or {}
        if stmt.get("title"):
            return stmt["title"]
    return stmt_key


# ──────────────────────────────────────────────────────────────────────────────
# 6) Excel export for consolidated statements
# ──────────────────────────────────────────────────────────────────────────────

def export_consolidated_excel(consolidated: dict, out_path: str) -> str:
    """
    Export the consolidated statements to a single Excel file.
    Sheets:
      - ملخص: group info, NCI breakdown, IC eliminations applied
      - قائمة: each statement (balance_sheet, income_statement, ...)
      - الإيضاحات: IC eliminations detail
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from pathlib import Path

    wb = Workbook()
    wb.remove(wb.active)

    bold_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E40AF")
    body_font = Font(name="Calibri", size=10)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    center = Alignment(horizontal="center", vertical="center", readingOrder=2)
    right = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # === Sheet 1: ملخص ===
    ws = wb.create_sheet("ملخص المجموعة")
    group = consolidated.get("group", {})
    companies = consolidated.get("companies", [])

    ws.cell(row=1, column=1, value=group.get("name", "مجموعة"))
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1); c.font = title_font; c.alignment = right
    ws.row_dimensions[1].height = 24

    ws.cell(row=2, column=1, value="الشركة الأم: " + (group.get("parent_company_name") or group.get("parent_company_id", "")))
    ws.merge_cells("A2:C2")
    ws.cell(row=2, column=1).alignment = right
    ws.cell(row=2, column=1).font = body_font

    # قائمة الشركات
    ws.cell(row=4, column=1, value="الشركات المدرجة في التجميع")
    ws.merge_cells("A4:C4")
    ws.cell(row=4, column=1).font = Font(name="Calibri", size=12, bold=True, color="1E40AF")
    ws.cell(row=4, column=1).alignment = right

    for col, h in enumerate(["الشركة", "الفترة", "نسبة الملكية"], 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = bold_white; cell.fill = header_fill; cell.alignment = center

    row = 6
    parent_id = group.get("parent_company_id")
    for jd, link in zip(_jobs_for_companies(consolidated), consolidated.get("ownership_links", [])):
        ws.cell(row=row, column=1, value=jd.get("company_name", "")).alignment = right
        ws.cell(row=row, column=2, value=jd.get("period", "")).alignment = center
        pct = link.get("ownership_pct", 0)
        ws.cell(row=row, column=3, value=f"{pct:.1f}%").alignment = center
        if link.get("company_id") == parent_id:
            ws.cell(row=row, column=3).font = Font(name="Calibri", size=10, bold=True, color="1E40AF")
        row += 1

    # NCI Breakdown
    row += 1
    ws.cell(row=row, column=1, value="الحصص غير المسيطر عليها (NCI)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=12, bold=True, color="92400E")
    row += 1
    nci = consolidated.get("nci", {})
    for col, (k, label) in enumerate([
        ("equity_total", "إجمالي حقوق الملكية المجمعة"),
        ("ownership_pct", "متوسط نسبة الملكية %"),
        ("parent_share", "حصة الشركة الأم"),
        ("nci", "الحصص غير المسيطر عليها (NCI)"),
    ], 1):
        ws.cell(row=row, column=col, value=label).font = body_font
        val = nci.get(k, 0)
        if k == "ownership_pct":
            ws.cell(row=row, column=col+1, value=f"{val:.1f}%").alignment = center
        else:
            ws.cell(row=row, column=col+1, value=val)
            ws.cell(row=row, column=col+1).number_format = "#,##0.00;(#,##0.00)"
            ws.cell(row=row, column=col+1).font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # IC Eliminations
    row += 1
    ws.cell(row=row, column=1, value="الاستبعادات البينية المطبقة")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=12, bold=True, color="15803D")
    row += 1
    ic = consolidated.get("ic_eliminations", {})
    ARABIC_LABELS = {
        "ic_receivable": "المدينون المتبادلون",
        "ic_payable": "الدائنون المتبادلون",
        "ic_revenue": "الإيرادات المتبادلة",
        "ic_expense": "المصاريف المتبادلة",
        "ic_loan_receivable": "قروض ممنوحة لشركات المجموعة",
        "ic_loan_payable": "قروض مستلمة من شركات المجموعة",
        "investment_in_sub": "استثمار الشركة الأم في التابعة",
        "ic_dividend_receivable": "توزيعات مستحقة من تابعة",
        "ic_dividend_payable": "توزيعات مستحقة لشركة أم",
    }
    for col, (k, label) in enumerate([
        ("ic_receivable", ARABIC_LABELS["ic_receivable"]),
        ("ic_payable", ARABIC_LABELS["ic_payable"]),
        ("ic_revenue", ARABIC_LABELS["ic_revenue"]),
        ("ic_expense", ARABIC_LABELS["ic_expense"]),
        ("investment_in_sub", ARABIC_LABELS["investment_in_sub"]),
    ], 1):
        ws.cell(row=row, column=col, value=label).font = body_font
        val = ic.get(k, 0)
        if val != 0:
            cell = ws.cell(row=row, column=col+1, value=val)
            cell.number_format = "#,##0.00;(#,##0.00)"
            cell.font = Font(name="Calibri", size=10, color="15803D", italic=True)
        else:
            ws.cell(row=row, column=col+1, value="—")
            ws.cell(row=row, column=col+1).font = Font(name="Calibri", size=10, color="9CA3AF")
        row += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # === Sheets for each statement ===
    for stmt_key, stmt in (consolidated.get("statements") or {}).items():
        sheet = wb.create_sheet(stmt_key[:31])
        sheet.cell(row=1, column=1, value=stmt.get("title", stmt_key))
        sheet.merge_cells("A1:B1")
        sheet.cell(row=1, column=1).font = title_font
        sheet.cell(row=1, column=1).alignment = right
        sheet.row_dimensions[1].height = 24

        for col, h in enumerate(["البند", "المبلغ الموحد"], 1):
            cell = sheet.cell(row=3, column=col, value=h)
            cell.font = bold_white; cell.fill = header_fill; cell.alignment = center

        r = 4
        for line in stmt.get("lines", []):
            sheet.cell(row=r, column=1, value=line.get("label", ""))
            sheet.cell(row=r, column=1).alignment = right
            cell = sheet.cell(row=r, column=2, value=float(line.get("amount", 0) or 0))
            cell.number_format = "#,##0.00;(#,##0.00)"
            cell.alignment = left
            if line.get("bold"):
                sheet.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True)
                cell.font = Font(name="Calibri", size=11, bold=True, color="1E40AF")
                sheet.cell(row=r, column=1).fill = total_fill
                cell.fill = total_fill
            if line.get("eliminated"):
                cell.font = Font(name="Calibri", size=10, color="15803D", italic=True)
            r += 1
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 22

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


def _jobs_for_companies(consolidated: dict) -> list[dict]:
    """Reconstruct a minimal company list from `companies` (display only)."""
    return consolidated.get("companies", [])
