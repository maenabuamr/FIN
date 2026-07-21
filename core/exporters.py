"""
Exporters — render financial statements and notes to Excel and PDF
with full Arabic support (no broken letters, no reversed words).
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
)

from .arabic_utils import ar, fmt_amount, clean
from .financial_statements import Statement
from .notes_generator import Note


# ──────────────────────────────────────────────────────────────────────────────
# Font registration for PDF (Arabic shaping)
# ──────────────────────────────────────────────────────────────────────────────

_ARABIC_FONT_NAME = "ArabicAmiri"
_ARABIC_FONT_BOLD = "ArabicAmiri-Bold"
_ARABIC_FONT_PATH = Path(__file__).parent.parent / "static" / "fonts" / "Amiri-Regular.ttf"
_ARABIC_FONT_BOLD_PATH = Path(__file__).parent.parent / "static" / "fonts" / "Amiri-Bold.ttf"


def _register_fonts() -> None:
    """Register Amiri (free, OFL) for Arabic shaping in PDF."""
    if _ARABIC_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return
    if _ARABIC_FONT_PATH.exists():
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_NAME, str(_ARABIC_FONT_PATH)))
    else:
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_NAME, "Helvetica"))
    if _ARABIC_FONT_BOLD_PATH.exists():
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_BOLD, str(_ARABIC_FONT_BOLD_PATH)))
    else:
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_BOLD, "Helvetica-Bold"))


# ──────────────────────────────────────────────────────────────────────────────
# Excel exporter
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(
    statements: dict[str, Statement],
    notes: list[Note],
    out_path: str,
    company_name: str = "الشركة",
    period: str = "",
) -> str:
    """Render all statements + notes into one .xlsx file."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Cover sheet
    cover = wb.create_sheet("ملخص")
    _excel_cover(cover, statements, company_name, period)

    # ── Each statement
    for key, stmt in statements.items():
        ws = wb.create_sheet(stmt.title[:31])
        _excel_statement(ws, stmt)

    # ── Notes
    if notes:
        ws = wb.create_sheet("الإيضاحات")
        _excel_notes(ws, notes)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


def _excel_cover(ws, statements, company_name, period) -> None:
    ws["A1"] = company_name
    ws["A1"].font = Font(name="Calibri", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="right", readingOrder=2)
    ws.merge_cells("A1:D1")

    ws["A2"] = period
    ws["A2"].font = Font(name="Calibri", size=12, italic=True)
    ws["A2"].alignment = Alignment(horizontal="right", readingOrder=2)
    ws.merge_cells("A2:D2")

    ws["A4"] = "القوائم المالية الرئيسية"
    ws["A4"].font = Font(name="Calibri", size=14, bold=True)
    ws["A4"].alignment = Alignment(horizontal="right", readingOrder=2)
    ws.merge_cells("A4:D4")

    headers = ["القائمة", "إجمالي الأصول / الإيرادات", "صافي الربح", "التوازن"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E293B")
        c.alignment = Alignment(horizontal="center", readingOrder=2)

    row = 6
    for key, stmt in statements.items():
        ws.cell(row=row, column=1, value=stmt.title)
        if "total_assets" in stmt.totals:
            ws.cell(row=row, column=2, value=stmt.totals["total_assets"])
        elif "total_revenue" in stmt.totals:
            ws.cell(row=row, column=2, value=stmt.totals["total_revenue"])
        ws.cell(row=row, column=3, value=stmt.totals.get("net_profit", "—"))
        ws.cell(row=row, column=4, value="✓ متوازن" if stmt.totals.get("balanced") else "—")
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", readingOrder=2
            )
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    # Number format for the value cells
    for r in range(6, 6 + len(statements)):
        for col in (2, 3):
            ws.cell(row=r, column=col).number_format = "#,##0.00"


def _excel_statement(ws, stmt: Statement) -> None:
    # Title
    ws["A1"] = stmt.title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
    ws.merge_cells("A1:C1")

    ws["A2"] = stmt.subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
    ws.merge_cells("A2:C2")

    ws["A3"] = stmt.as_of or stmt.period
    ws["A3"].font = Font(name="Calibri", size=10, color="475569")
    ws["A3"].alignment = Alignment(horizontal="center", readingOrder=2)
    ws.merge_cells("A3:C3")

    # Header row
    headers = ["البيان", stmt.currency, "إيضاح"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E293B")
        c.alignment = Alignment(horizontal="center", readingOrder=2)
        c.border = _excel_border()

    thin = _excel_border()
    row = 6
    for line in stmt.lines:
        # Indent in label
        prefix = ("    " * line.indent)
        ws.cell(row=row, column=1, value=prefix + line.label)
        ws.cell(row=row, column=2, value=line.amount if abs(line.amount) > 1e-9 else "—")
        ws.cell(row=row, column=3, value=line.ref)
        # Style
        if line.is_total or line.bold:
            font = Font(name="Calibri", size=11, bold=True)
        else:
            font = Font(name="Calibri", size=11)
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = Alignment(horizontal="center" if col == 2 else "right", readingOrder=2)
            cell.border = thin
        if line.is_subtotal or line.is_total:
            fill = PatternFill("solid", fgColor="F1F5F9")
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = fill
        # Number format
        ws.cell(row=row, column=2).number_format = '#,##0.00;(#,##0.00);"—"'
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10


def _excel_notes(ws, notes: list[Note]) -> None:
    ws["A1"] = "الإيضاحات المرفقة مع القوائم المالية"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="right", readingOrder=2)
    ws.merge_cells("A1:D1")

    row = 3
    for n in notes:
        ws.cell(row=row, column=1, value=f"{n.number} - {n.title}")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", readingOrder=2)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        ws.cell(row=row, column=1, value=n.body)
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, color="334155")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", readingOrder=2, wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 40
        row += 1

        if n.accounts:
            # Sub-header
            for i, h in enumerate(["الحساب", "الرمز", "المبلغ", ""], start=1):
                c = ws.cell(row=row, column=i, value=h)
                c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1E293B")
                c.alignment = Alignment(horizontal="center", readingOrder=2)
            row += 1
            for a in n.accounts:
                ws.cell(row=row, column=1, value=a["name"])
                ws.cell(row=row, column=2, value=a["code"])
                ws.cell(row=row, column=3, value=a["amount"] if abs(a["amount"]) > 1e-9 else "—")
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(name="Calibri", size=11)
                    cell.alignment = Alignment(horizontal="center" if col in (2, 3) else "right", readingOrder=2)
                ws.cell(row=row, column=3).number_format = '#,##0.00;(#,##0.00)'
                row += 1
            row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 10


def _excel_border() -> Border:
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


# ──────────────────────────────────────────────────────────────────────────────
# PDF exporter
# ──────────────────────────────────────────────────────────────────────────────

def export_pdf(
    statements: dict[str, Statement],
    notes: list[Note],
    out_path: str,
    company_name: str = "الشركة",
    period: str = "",
) -> str:
    """Render all statements + notes into a multi-page PDF."""
    _register_fonts()

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(out),
        pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=company_name,
        author="Financial Review System",
    )

    styles = _pdf_styles()

    story: list = []
    # Cover
    story.append(Paragraph(ar(company_name), styles["cover_title"]))
    if period:
        story.append(Paragraph(ar(period), styles["cover_subtitle"]))
    story.append(Spacer(1, 1.5 * cm))
    story.append(Paragraph(ar("القوائم المالية الرئيسية"), styles["h1"]))
    for key, stmt in statements.items():
        story.append(Paragraph(ar(f"• {stmt.title}"), styles["li"]))
    story.append(PageBreak())

    # Each statement
    for key, stmt in statements.items():
        story.extend(_statement_flowables(stmt, styles))
        story.append(PageBreak())

    # Notes
    if notes:
        story.append(Paragraph(ar("الإيضاحات المرفقة"), styles["h1"]))
        for n in notes:
            story.append(Paragraph(ar(f"إيضاح ({n.number}) — {n.title}"), styles["h2"]))
            story.append(Paragraph(ar(n.body), styles["body"]))
            if n.accounts:
                story.append(_notes_table(n))
            story.append(Spacer(1, 0.4 * cm))

    doc.build(story)
    return str(out)


def _pdf_styles() -> dict:
    return {
        "cover_title": ParagraphStyle(
            "CoverTitle", fontName=_ARABIC_FONT_BOLD, fontSize=28,
            alignment=2, textColor=colors.HexColor("#0F172A"),
            spaceAfter=12, leading=34,
        ),
        "cover_subtitle": ParagraphStyle(
            "CoverSub", fontName=_ARABIC_FONT_NAME, fontSize=14,
            alignment=2, textColor=colors.HexColor("#475569"),
            spaceAfter=20, leading=20,
        ),
        "h1": ParagraphStyle(
            "H1", fontName=_ARABIC_FONT_BOLD, fontSize=18,
            alignment=2, textColor=colors.HexColor("#0F172A"),
            spaceAfter=10, leading=24,
        ),
        "h2": ParagraphStyle(
            "H2", fontName=_ARABIC_FONT_BOLD, fontSize=14,
            alignment=2, textColor=colors.HexColor("#0F172A"),
            spaceAfter=8, leading=20,
        ),
        "body": ParagraphStyle(
            "Body", fontName=_ARABIC_FONT_NAME, fontSize=11,
            alignment=2, textColor=colors.HexColor("#334155"),
            spaceAfter=6, leading=18,
        ),
        "li": ParagraphStyle(
            "Li", fontName=_ARABIC_FONT_NAME, fontSize=12,
            alignment=2, textColor=colors.HexColor("#334155"),
            spaceAfter=4, leading=18,
        ),
        "cell_label": ParagraphStyle(
            "CellLabel", fontName=_ARABIC_FONT_NAME, fontSize=10,
            alignment=2, leading=14,
        ),
        "cell_amount": ParagraphStyle(
            "CellAmount", fontName=_ARABIC_FONT_NAME, fontSize=10,
            alignment=1, leading=14,
        ),
        "cell_total": ParagraphStyle(
            "CellTotal", fontName=_ARABIC_FONT_BOLD, fontSize=11,
            alignment=2, leading=14,
        ),
    }


def _statement_flowables(stmt: Statement, styles: dict) -> list:
    out: list = []
    out.append(Paragraph(ar(stmt.title), styles["h1"]))
    out.append(Paragraph(ar(stmt.subtitle), styles["li"]))
    if stmt.as_of:
        out.append(Paragraph(ar(f"كما في {stmt.as_of}"), styles["body"]))
    if stmt.period:
        out.append(Paragraph(ar(f"عن الفترة: {stmt.period}"), styles["body"]))
    out.append(Spacer(1, 0.4 * cm))

    # Build the table
    data = []
    for line in stmt.lines:
        if abs(line.amount) > 1e-9 or line.is_subtotal or line.is_total or line.is_total:
            indent = "&nbsp;" * (line.indent * 4)
            label = ar(f"{indent}{line.label}")
            amount = fmt_amount(line.amount) if abs(line.amount) > 1e-9 else "—"
            ref = ar(line.ref) if line.ref else ""
            data.append([label, amount, ref])
        else:
            data.append([ar(line.label), "", ""])

    if not data:
        return out

    t = Table(data, colWidths=[10 * cm, 4.5 * cm, 1.5 * cm])
    t.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONT", (0, 0), (-1, -1), _ARABIC_FONT_NAME, 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
    ]))
    # Bold & shade subtotals/totals
    for i, line in enumerate(stmt.lines):
        if line.is_subtotal:
            t.setStyle(TableStyle([
                ("FONT", (0, i), (-1, i), _ARABIC_FONT_BOLD, 10),
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F1F5F9")),
            ]))
        if line.is_total:
            t.setStyle(TableStyle([
                ("FONT", (0, i), (-1, i), _ARABIC_FONT_BOLD, 11),
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E2E8F0")),
                ("LINEABOVE", (0, i), (-1, i), 0.8, colors.HexColor("#0F172A")),
            ]))
    out.append(t)
    return out


def _notes_table(n: Note) -> Table:
    data = [[ar("الحساب"), ar("الرمز"), ar("المبلغ")]]
    for a in n.accounts:
        data.append([
            ar(a["name"]),
            ar(a["code"]),
            fmt_amount(a["amount"]),
        ])
    t = Table(data, colWidths=[10 * cm, 3 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONT", (0, 0), (-1, -1), _ARABIC_FONT_NAME, 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), _ARABIC_FONT_BOLD, 10),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


# ──────────────────────────────────────────────────────────────────────────────
# Comparison exporter — Excel/PDF for two periods
# ──────────────────────────────────────────────────────────────────────────────

def export_comparison_excel(
    comparisons: dict,   # {"balance_sheet": [...], "income_statement": [...], ...}
    kpis: list[dict],
    out_path: str,
    company_name: str = "الشركة",
    period_current: str = "",
    period_prior: str = "",
    detailed_notes: list[dict] | None = None,  # rows of {num, title, period, total, body, accounts}
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("ملخص")
    cover["A1"] = f"{company_name} - مقارنة الفترات"
    cover["A1"].font = Font(name="Calibri", size=20, bold=True)
    cover["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
    cover.merge_cells("A1:C1")
    cover["A2"] = f"{period_prior}  ←→  {period_current}"
    cover["A2"].font = Font(name="Calibri", size=12, italic=True)
    cover["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
    cover.merge_cells("A2:C2")

    # KPI table
    cover["A4"] = "المؤشرات الرئيسية"
    cover["A4"].font = Font(name="Calibri", size=14, bold=True)
    cover["A4"].alignment = Alignment(horizontal="right", readingOrder=2)
    cover.merge_cells("A4:C4")
    headers = ["المؤشر", "الفترة الحالية", "الفترة السابقة"]
    for i, h in enumerate(headers, start=1):
        c = cover.cell(row=5, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E293B")
        c.alignment = Alignment(horizontal="center", readingOrder=2)
    row = 6
    for k in kpis:
        cover.cell(row=row, column=1, value=k["name"])
        cover.cell(row=row, column=2, value=k["current"])
        cover.cell(row=row, column=3, value=k["prior"])
        for col in range(1, 4):
            cover.cell(row=row, column=col).alignment = Alignment(
                horizontal="center" if col != 1 else "right", readingOrder=2
            )
        cover.cell(row=row, column=2).number_format = "#,##0.00"
        cover.cell(row=row, column=3).number_format = "#,##0.00"
        row += 1
    for col, w in zip("ABC", [40, 22, 22]):
        cover.column_dimensions[col].width = w

    # Per-statement comparison sheets
    titles = {
        "balance_sheet": "المركز المالي - مقارنة",
        "income_statement": "الدخل - مقارنة",
        "cash_flow": "التدفقات النقدية - مقارنة",
        "equity": "حقوق الملكية - مقارنة",
        "__notes__": "الإيضاحات - مقارنة",
    }
    for key, rows in comparisons.items():
        ws = wb.create_sheet(titles.get(key, key)[:31])
        ws["A1"] = titles.get(key, key)
        ws["A1"].font = Font(name="Calibri", size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
        ws.merge_cells("A1:C1")

        hdr = ["البند", "الفترة الحالية", "الفترة السابقة"]
        for i, h in enumerate(hdr, start=1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1E293B")
            c.alignment = Alignment(horizontal="center", readingOrder=2)

        for i, line in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=line["label"])
            ws.cell(row=i, column=2, value=line["current"])
            ws.cell(row=i, column=3, value=line["prior"])
            font = Font(name="Calibri", size=11, bold=line.get("bold", False))
            for col in range(1, 4):
                c = ws.cell(row=i, column=col)
                c.font = font
                c.alignment = Alignment(horizontal="center" if col != 1 else "right", readingOrder=2)
            ws.cell(row=i, column=2).number_format = "#,##0.00;(#,##0.00)"
            ws.cell(row=i, column=3).number_format = "#,##0.00;(#,##0.00)"
        for col, w in zip("ABC", [45, 22, 22]):
            ws.column_dimensions[col].width = w

    # Detailed notes comparison sheet
    if detailed_notes:
        wn = wb.create_sheet("الإيضاحات - تفاصيل مقارنة")
        wn["A1"] = "الإيضاحات - تفاصيل المقارنة"
        wn["A1"].font = Font(name="Calibri", size=16, bold=True)
        wn["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
        wn.merge_cells("A1:F1")
        wn["A2"] = f"مقارنة {period_prior} ←→ {period_current}"
        wn["A2"].font = Font(name="Calibri", size=11, italic=True, color="6B7280")
        wn["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
        wn.merge_cells("A2:F2")
        hdrs = ["رقم", "عنوان الإيضاح", "الفترة", "الرصيد", "الوصف", "الحسابات"]
        for i, h in enumerate(hdrs, start=1):
            cell = wn.cell(row=4, column=i, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.alignment = Alignment(horizontal="center", readingOrder=2)
        for i, row in enumerate(detailed_notes, start=5):
            wn.cell(row=i, column=1, value=row.get("num", ""))
            wn.cell(row=i, column=2, value=row.get("title", ""))
            wn.cell(row=i, column=3, value=row.get("period", ""))
            wn.cell(row=i, column=4, value=row.get("total", 0))
            wn.cell(row=i, column=5, value=row.get("body", ""))
            wn.cell(row=i, column=6, value=row.get("accounts", ""))
            for col in range(1, 7):
                cell = wn.cell(row=i, column=col)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(
                    horizontal="right" if col in (1,2,3,5,6) else "left", readingOrder=2, wrap_text=True
                )
                if col == 4:
                    cell.number_format = "#,##0.00;(#,##0.00)"
                    if (row.get("total") or 0) < 0:
                        cell.font = Font(name="Calibri", size=10, color="DC2626")
            # group alternating by serial number
            num = row.get("num", 0)
            if isinstance(num, int) and num % 2 == 0:
                for col in range(1, 7):
                    wn.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F8FAFC")
        for col, w in zip("ABCDEF", [6, 35, 16, 18, 55, 60]):
            wn.column_dimensions[col].width = w
        wn.row_dimensions[1].height = 28
        wn.row_dimensions[4].height = 24

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


def export_notes_comparison_sheet(wb, detailed_notes: list[dict], period_current: str = "", period_prior: str = ""):
    """
    بناء ورقة "الإيضاحات المرفقة" في workbook موجود.
    كل إيضاح يأخذ قسم كامل بجدول واحد 5 أعمدة:
      - الحساب | الرمز | الفترة الحالية | الفترة السابقة | الفرق
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("الإيضاحات المرفقة")

    bold_white = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E40AF")
    body_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    current_header_fill = PatternFill("solid", fgColor="1E40AF")
    prev_header_fill = PatternFill("solid", fgColor="92400E")
    diff_header_fill = PatternFill("solid", fgColor="15803D")
    total_fill = PatternFill("solid", fgColor="F1F5F9")
    diff_row_fill = PatternFill("solid", fgColor="1E40AF")
    center = Alignment(horizontal="center", vertical="center", readingOrder=2, wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # عرض الأعمدة: الحساب | الرمز | الحالية | السابقة | الفرق
    col_widths = [55, 18, 22, 22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    row = 1
    for note in detailed_notes:
        # 1) عنوان الإيضاح: "1 - النقدية وما في حكمها"
        ws.cell(row=row, column=1, value=f'{note.get("number", "")} - {note.get("title", "")}')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1)
        cell.font = title_font
        cell.alignment = right
        cell.fill = total_fill
        ws.row_dimensions[row].height = 24
        row += 1

        # 2) الوصف
        body = note.get("body", "")
        if body:
            ws.cell(row=row, column=1, value=body)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1)
            cell.font = body_font
            cell.alignment = right
            ws.row_dimensions[row].height = 32
            row += 1

        # 3) Headers: الحساب | الرمز | الحالية | السابقة | الفرق
        headers = [
            ("الحساب", header_fill),
            ("الرمز", header_fill),
            (f"الفترة الحالية ({period_current})" if period_current else "الفترة الحالية", current_header_fill),
            (f"الفترة السابقة ({period_prior})" if period_prior else "الفترة السابقة", prev_header_fill),
        ]
        for col, (h, fill) in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold_white
            cell.fill = fill
            cell.alignment = center
        ws.row_dimensions[row].height = 28
        row += 1

        # 4) بناء قاموس من الكود إلى المبلغ للفترة السابقة للمطابقة
        cur_accounts = note.get("current_accounts", [])
        prev_accounts = note.get("previous_accounts", [])
        prev_map = {str(a.get("code", "")): (a.get("amount", 0) or 0) for a in prev_accounts}
        cur_map = {str(a.get("code", "")): (a.get("amount", 0) or 0) for a in cur_accounts}

        # جمع كل الأكواد (من الفترة الحالية + السابقة)
        all_codes = []
        seen = set()
        for a in cur_accounts:
            code = str(a.get("code", ""))
            if code not in seen:
                seen.add(code); all_codes.append((code, a, 'cur'))
        for a in prev_accounts:
            code = str(a.get("code", ""))
            if code not in seen:
                seen.add(code); all_codes.append((code, a, 'prev'))

        # ترتيب حسب المبلغ تنازلياً (الأكبر أولاً)
        def _sort_key(item):
            code, a, src = item
            return -(max(abs(cur_map.get(code, 0)), abs(prev_map.get(code, 0))))
        all_codes.sort(key=_sort_key)

        for code, a, src in all_codes:
            name = a.get("name", "")
            cur_amt = cur_map.get(code, 0)
            prev_amt = prev_map.get(code, 0)
            diff_amt = cur_amt - prev_amt

            ws.cell(row=row, column=1, value=name).alignment = right
            ws.cell(row=row, column=2, value=code).alignment = Alignment(horizontal="center", readingOrder=2)
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, color="475569")

            for col_idx, val in [(3, cur_amt), (4, prev_amt)]:
                c = ws.cell(row=row, column=col_idx, value=val)
                c.number_format = "#,##0.00;(#,##0.00)"
                c.alignment = left
                if val < 0:
                    c.font = Font(name="Calibri", size=10, color="DC2626", bold=(col_idx == 5))
                else:
                    c.font = Font(name="Calibri", size=10, color="0F172A", bold=(col_idx == 5))
            row += 1

        if not all_codes:
            ws.cell(row=row, column=1, value="— لا توجد حسابات —")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).font = body_font
            row += 1

        # 5) صف المجموع
        ws.cell(row=row, column=1, value="المجموع")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=1).alignment = right
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, color="1E40AF", bold=True)
        for col_idx, val, color in [
            (3, note.get("current_total", 0), "1E40AF"),
            (4, note.get("previous_total", 0), "92400E"),
        ]:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = Font(name="Calibri", size=11, bold=True, color=color)
            c.number_format = "#,##0.00;(#,##0.00)"
            c.alignment = left
            c.fill = total_fill
        row += 2  # سطر فاصل

    return ws
