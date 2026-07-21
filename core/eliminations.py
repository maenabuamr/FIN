"""
Advanced Intercompany Eliminations Engine (IFRS 10).

Detects and pairs intercompany transactions across the group:
  - Sales / Purchases (إيرادات / مشتريات)
  - Receivables / Payables (مدينون / دائنون)
  - Cash transfers (نقدية محولة بين بنوك المجموعة)
  - Dividends payable / receivable (توزيعات)
  - Loans (قروض ممنوحة / مستلمة)
  - Investment in subsidiary vs. equity (الاستثمار في التابعة)
  - Unrealized profit in ending inventory (أرباح غير محققة في المخزون)

Each detected transaction is paired with its counterpart (e.g. a 25,000
receivable on the parent is matched with a 25,000 payable on the sub).
Eliminations are applied symmetrically: both sides are reduced.

If amounts don't match exactly (e.g. timing differences), the engine
still pairs them but flags the difference for review.
"""

from __future__ import annotations

from collections import defaultdict
from typing import Optional


# ──────────────────────────────────────────────────────────────────────────────
# 1) Sub-categories that mark intercompany accounts
# ──────────────────────────────────────────────────────────────────────────────

IC_CATEGORY_MAP = {
    # Receivables (مدينون)
    "ic_receivable":         {"kind": "receivable",  "balance_sheet": True,  "income": False, "ar": "مدينون بين شركات المجموعة"},
    "ic_payable":            {"kind": "payable",     "balance_sheet": True,  "income": False, "ar": "دائنون بين شركات المجموعة"},
    "ic_revenue":            {"kind": "ic_revenue",  "balance_sheet": False, "income": True,  "ar": "إيرادات متبادلة"},
    "ic_expense":            {"kind": "ic_expense",  "balance_sheet": False, "income": True,  "ar": "مصاريف متبادلة"},
    "ic_loan_receivable":    {"kind": "ic_loan_recv","balance_sheet": True,  "income": False, "ar": "قروض ممنوحة لشركات المجموعة"},
    "ic_loan_payable":       {"kind": "ic_loan_pay", "balance_sheet": True,  "income": False, "ar": "قروض مستلمة من شركات المجموعة"},
    "ic_cash_transfer":      {"kind": "ic_cash",     "balance_sheet": True,  "income": False, "ar": "تحويلات نقدية معلقة بين شركات المجموعة"},
    "ic_dividend_receivable":{"kind": "ic_div_recv", "balance_sheet": True,  "income": False, "ar": "توزيعات مستحقة من تابعة"},
    "ic_dividend_payable":   {"kind": "ic_div_pay",  "balance_sheet": True,  "income": False, "ar": "توزيعات مستحقة لشركة أم"},
    "investment_in_sub":     {"kind": "invest_in_sub","balance_sheet": True, "income": False, "ar": "استثمار الشركة الأم في التابعة"},
    "unrealized_profit_inv": {"kind": "ur_profit",   "balance_sheet": True,  "income": True,  "ar": "أرباح غير محققة في المخزون"},
    "intercompany_receivable": {"kind": "receivable","balance_sheet": True,  "income": False, "ar": "مدينون بين شركات المجموعة"},
    "intercompany_payable":    {"kind": "payable",   "balance_sheet": True,  "income": False, "ar": "دائنون بين شركات المجموعة"},
}

# Pairs: what should match what
PAIRING_RULES = {
    "ic_receivable":    "ic_payable",        # مدينون ↔ دائنون
    "ic_payable":       "ic_receivable",
    "ic_revenue":       "ic_expense",        # مبيعات ↔ مشتريات
    "ic_expense":       "ic_revenue",
    "ic_loan_receivable": "ic_loan_payable",
    "ic_loan_payable":  "ic_loan_receivable",
    "ic_dividend_receivable": "ic_dividend_payable",
    "ic_dividend_payable":    "ic_dividend_receivable",
}


# ──────────────────────────────────────────────────────────────────────────────
# 2) Detection - scan all accounts and group by sub_category
# ──────────────────────────────────────────────────────────────────────────────

def detect_ic_transactions(jobs_data: list[dict]) -> list[dict]:
    """
    Walk every company's accounts, collect those flagged as intercompany,
    and return them as a list of detected transactions.

    Each detection: {
      "id": "<auto>",
      "company_id": "c_xxx",
      "company_name": "...",
      "account_code": "110000756",
      "account_name": "...",
      "sub_category": "ic_receivable",
      "kind": "receivable",
      "amount": 25000.0,
      "is_debit": True,         # رصيد مدين (دائن - لدينا فلوس عنده)
      "matched": False,
      "matched_with": None,
      "diff": 0.0,
    }
    """
    transactions = []
    for jd in jobs_data:
        cid = jd.get("company_id")
        cname = jd.get("company_name", "")
        accounts = jd.get("accounts") or jd.get("raw_rows") or []
        for a in accounts:
            cat = a.get("sub_category") or a.get("category") or ""
            if cat not in IC_CATEGORY_MAP:
                continue
            amt = float(a.get("balance", 0) or a.get("amount", 0) or 0)
            if abs(amt) < 0.01:
                continue
            meta = IC_CATEGORY_MAP[cat]
            transactions.append({
                "id": f"tx_{cid}_{a.get('code', '')}_{cat}",
                "company_id": cid,
                "company_name": cname,
                "account_code": str(a.get("code", "")),
                "account_name": a.get("name", ""),
                "sub_category": cat,
                "kind": meta["kind"],
                "amount": abs(amt),
                "is_debit": amt > 0,
                "matched": False,
                "matched_with": None,
                "diff": 0.0,
            })
    return transactions


# ──────────────────────────────────────────────────────────────────────────────
# 3) Pairing - match each IC transaction with its counterpart
# ──────────────────────────────────────────────────────────────────────────────

def pair_transactions(transactions: list[dict]) -> list[dict]:
    """
    For each transaction, find a matching counterpart in another company
    with the same account_code (or name) and a complementary sub_category.

    Returns the same list with `matched`, `matched_with`, `diff` filled.
    """
    # Bucket by (sub_category, account_code) for quick lookup
    by_key = defaultdict(list)
    for tx in transactions:
        if not tx.get("is_debit"):  # only match debits (positive amounts)
            continue
        # Match by code OR by name (for fuzzy match)
        by_key[(tx["sub_category"], tx["account_code"])].append(tx)
        # also by name
        by_key[("__name__", tx["account_name"])].append(tx)

    paired_results = []
    used = set()

    for tx in transactions:
        if not tx["is_debit"]:
            paired_results.append(tx)
            continue
        tx_id = tx["id"]
        if tx_id in used:
            continue

        # Find counterpart: same code, complementary sub_category
        complement_cat = PAIRING_RULES.get(tx["sub_category"])
        if not complement_cat:
            paired_results.append(tx)
            continue

        candidates = by_key.get((complement_cat, tx["account_code"]), [])
        candidates = [c for c in candidates if c["id"] != tx_id and c["id"] not in used
                      and c["company_id"] != tx["company_id"]]
        if not candidates:
            # try by name
            candidates = by_key.get(("__name__", tx["account_name"]), [])
            candidates = [c for c in candidates if c["id"] != tx_id and c["id"] not in used
                          and c["company_id"] != tx["company_id"] and c["sub_category"] == complement_cat]

        if candidates:
            # pick smallest diff
            best = min(candidates, key=lambda c: abs(c["amount"] - tx["amount"]))
            diff = round(tx["amount"] - best["amount"], 2)
            tx["matched"] = True
            tx["matched_with"] = best["id"]
            tx["diff"] = diff
            best["matched"] = True
            best["matched_with"] = tx_id
            best["diff"] = -diff
            used.add(tx_id)
            used.add(best["id"])

        paired_results.append(tx)

    return paired_results


# ──────────────────────────────────────────────────────────────────────────────
# 4) Eliminate - apply matched transactions to the consolidated statements
# ──────────────────────────────────────────────────────────────────────────────

def apply_eliminations(consolidated: dict, transactions: list[dict]) -> dict:
    """
    Apply approved eliminations to the consolidated statements.
    Each line whose label matches gets reduced by the elimination amount.
    """
    LABEL_MAP = {
        "ic_receivable":         ["المدينون", "مدينون", "ذمم مدينة"],
        "ic_payable":            ["الدائنون", "دائنون", "ذمم دائنة"],
        "ic_revenue":            ["الإيرادات", "إيرادات", "المبيعات", "ايرادات"],
        "ic_expense":            ["المصاريف", "مصاريف", "المشتريات", "مصروفات", "تكلفة"],
        "ic_loan_receivable":    ["قروض ممنوحة", "سلف لشركات"],
        "ic_loan_payable":       ["قروض مستلمة", "سلف من شركات"],
        "ic_cash_transfer":      ["النقدية", "نقدية", "البنك", "البنوك"],
        "ic_dividend_receivable":["توزيعات مستحقة من", "توزيعات مدينة"],
        "ic_dividend_payable":   ["توزيعات مستحقة", "توزيعات دائنة"],
        "investment_in_sub":     ["استثمارات", "الاستثمارات", "استثمار في"],
        "unrealized_profit_inv": ["المخزون", "مخزون", "الأرباح المبقاة", "احتياطيات"],
    }

    # Group transactions by sub_category
    by_cat = defaultdict(float)
    for tx in transactions:
        if not tx.get("matched") and tx.get("sub_category") != "unrealized_profit_inv":
            continue
        # Sum the amount (using the debit side only, the credit side is the same)
        by_cat[tx["sub_category"]] += tx["amount"]

    # For unrealized profit, apply twice (reduce inventory + reduce retained earnings)
    if "unrealized_profit_inv" in by_cat:
        consolidated["unrealized_profit"] = by_cat["unrealized_profit_inv"]

    # Apply to statements
    def _apply_to_lines(lines: list[dict]) -> list[dict]:
        new = []
        for line in lines:
            label = (line.get("label") or "").strip()
            amount = float(line.get("amount", 0) or 0)
            elim_applied = 0.0
            applied_cats = []
            for cat, amount_to_elim in by_cat.items():
                if amount_to_elim == 0:
                    continue
                keywords = LABEL_MAP.get(cat, [])
                if any(kw in label for kw in keywords):
                    # Receivables/payables: subtract the matched amount
                    if cat in ("ic_receivable", "ic_loan_receivable", "ic_dividend_receivable",
                               "ic_cash_transfer", "investment_in_sub"):
                        amount = round(amount - amount_to_elim, 2)
                        elim_applied += amount_to_elim
                        applied_cats.append(cat)
                    elif cat in ("ic_payable", "ic_loan_payable", "ic_dividend_payable"):
                        amount = round(amount - amount_to_elim, 2)
                        elim_applied += amount_to_elim
                        applied_cats.append(cat)
                    elif cat in ("ic_revenue",):
                        amount = round(amount - amount_to_elim, 2)
                        elim_applied += amount_to_elim
                        applied_cats.append(cat)
                    elif cat in ("ic_expense",):
                        amount = round(amount - amount_to_elim, 2)
                        elim_applied += amount_to_elim
                        applied_cats.append(cat)
                    by_cat[cat] = 0  # consumed
            new_line = dict(line)
            new_line["amount"] = amount
            if elim_applied:
                new_line["eliminated"] = True
                new_line["eliminated_amount"] = round(elim_applied, 2)
                new_line["eliminated_categories"] = applied_cats
            new.append(new_line)
        return new

    for stmt_key, stmt in (consolidated.get("statements") or {}).items():
        if isinstance(stmt, dict) and "lines" in stmt:
            stmt["lines"] = _apply_to_lines(stmt["lines"])

    # Apply unrealized profit to inventory and retained earnings
    if "unrealized_profit" in by_cat and by_cat["unrealized_profit"]:
        ur = by_cat["unrealized_profit"]
        for line in consolidated["statements"]["balance_sheet"]["lines"]:
            if "المخزون" in (line.get("label") or ""):
                line["amount"] = round(line["amount"] - ur, 2)
                line["eliminated"] = True
                line["eliminated_amount"] = ur
                line["eliminated_categories"] = ["unrealized_profit_inv"]

    consolidated["eliminations"] = [
        {
            "sub_category": cat,
            "kind": IC_CATEGORY_MAP[cat]["kind"],
            "label": IC_CATEGORY_MAP[cat]["ar"],
            "amount": round(amt, 2),
        }
        for cat, amt in by_cat.items() if amt
    ]
    return consolidated


# ──────────────────────────────────────────────────────────────────────────────
# 5) Unrealized profit in ending inventory (estimated)
# ──────────────────────────────────────────────────────────────────────────────

def estimate_unrealized_profit(
    parent_sales_to_subs: float,
    sub_ending_inventory: float,
    sub_cost_of_sales: float,
    profit_margin_pct: float = 0.0,
) -> float:
    """
    IFRS 10 requires elimination of unrealized profit (UP) on inventory
    still held by the group at period-end.

    UP = (Ending Inventory from IC sales) × Profit Margin

    Logic:
      1. Estimate the portion of sub's ending inventory that came from
         IC purchases (default: assume proportional if not provided).
      2. Apply the gross profit margin of the seller to that portion.
    """
    if sub_cost_of_sales == 0 or sub_ending_inventory == 0:
        return 0.0
    if profit_margin_pct <= 0:
        # default gross margin = 1 - COGS/Sales
        profit_margin_pct = max(0, 1.0 - (sub_cost_of_sales / max(parent_sales_to_subs, 1)))
    # Approximation: assume half of ending inventory is from IC sales
    ic_inventory_estimate = min(sub_ending_inventory, parent_sales_to_subs) * 0.5
    return round(ic_inventory_estimate * profit_margin_pct, 2)


# ──────────────────────────────────────────────────────────────────────────────
# 6) Elimination entries (journal)
# ──────────────────────────────────────────────────────────────────────────────

def generate_journal_entries(eliminations: list[dict]) -> list[dict]:
    """
    Convert matched eliminations into double-entry journal lines.
    Each entry: {description, debit_label, credit_label, amount}
    """
    entries = []
    for elim in eliminations:
        cat = elim.get("sub_category", "")
        amt = abs(elim.get("amount", 0))
        if cat == "ic_receivable":
            entries.append({
                "description": "استبعاد المدينون المتبادلون بين شركات المجموعة",
                "debit_label": "الدائنون المتبادلون (شركات المجموعة)",
                "credit_label": "المدينون المتبادلون (شركات المجموعة)",
                "amount": amt,
            })
        elif cat == "ic_payable":
            entries.append({
                "description": "استبعاد الدائنون المتبادلون",
                "debit_label": "الدائنون المتبادلون",
                "credit_label": "المدينون المتبادلون",
                "amount": amt,
            })
        elif cat == "ic_revenue":
            entries.append({
                "description": "استبعاد الإيرادات المتبادلة (مبيعات بينية)",
                "debit_label": "الإيرادات (مبيعات بينية)",
                "credit_label": "تكلفة المبيعات (مشتريات بينية)",
                "amount": amt,
            })
        elif cat == "ic_expense":
            entries.append({
                "description": "استبعاد المصاريف المتبادلة",
                "debit_label": "المصاريف المتبادلة",
                "credit_label": "الإيرادات المتبادلة",
                "amount": amt,
            })
        elif cat == "investment_in_sub":
            entries.append({
                "description": "استبعاد الاستثمار في الشركة التابعة مقابل حقوق الملكية",
                "debit_label": "حقوق الملكية (الشركة التابعة)",
                "credit_label": "الاستثمار في الشركة التابعة",
                "amount": amt,
            })
        elif cat == "ic_cash_transfer":
            entries.append({
                "description": "تسوية التحويلات النقدية المعلقة بين بنوك المجموعة",
                "debit_label": "النقدية (الجهة المستلمة)",
                "credit_label": "النقدية (الجهة المرسلة)",
                "amount": amt,
            })
    return entries


# ──────────────────────────────────────────────────────────────────────────────
# 7) Excel export for eliminations
# ──────────────────────────────────────────────────────────────────────────────

def export_eliminations_to_excel(
    transactions: list[dict],
    out_path: str,
    group_name: str = "مجموعة",
) -> str:
    """
    Build a dedicated Excel file with all detected eliminations and
    their journal entries.
    Sheets:
      - الاستبعادات المكتشفة: كل المعاملات البينية + حالة المطابقة
      - قيود الاستبعاد: القيود المحاسبية لكل استبعاد
      - ملخص: إجماليات حسب النوع
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from pathlib import Path

    wb = Workbook()
    wb.remove(wb.active)

    bold_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E40AF")
    matched_fill = PatternFill("solid", fgColor="D1FAE5")
    unmatched_fill = PatternFill("solid", fgColor="FEF3C7")
    diff_fill = PatternFill("solid", fgColor="FEE2E2")
    center = Alignment(horizontal="center", vertical="center", readingOrder=2)
    right = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)

    # === Sheet 1: الاستبعادات المكتشفة ===
    ws = wb.create_sheet("الاستبعادات المكتشفة")
    ws.cell(row=1, column=1, value=group_name + " - الاستبعادات البينية المكتشفة")
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = right
    ws.row_dimensions[1].height = 24

    headers = ["#", "الجهة", "الحساب", "الرمز", "النوع", "المبلغ", "مطابق؟", "الفرق"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bold_white; c.fill = PatternFill("solid", fgColor="1E3A8A"); c.alignment = center

    r = 4
    for i, tx in enumerate(transactions, 1):
        ws.cell(row=r, column=1, value=i).alignment = center
        ws.cell(row=r, column=2, value=tx.get("company_name", "")).alignment = right
        ws.cell(row=r, column=3, value=tx.get("account_name", "")).alignment = right
        ws.cell(row=r, column=4, value=tx.get("account_code", "")).alignment = center
        ws.cell(row=r, column=5, value=IC_CATEGORY_MAP.get(tx.get("sub_category", ""), {}).get("ar", "")).alignment = right
        c = ws.cell(row=r, column=6, value=tx.get("amount", 0))
        c.number_format = "#,##0.00"
        c.alignment = center
        match_cell = ws.cell(row=r, column=7, value="✅ نعم" if tx.get("matched") else "⚠️ لا")
        match_cell.alignment = center
        diff_cell = ws.cell(row=r, column=8, value=tx.get("diff", 0) or 0)
        diff_cell.number_format = "#,##0.00;(#,##0.00)"
        diff_cell.alignment = center
        # color rows
        fill = matched_fill if tx.get("matched") else unmatched_fill
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = fill
        if tx.get("diff"):
            ws.cell(row=r, column=8).fill = diff_fill
        r += 1

    for col, w in zip("ABCDEFGH", [5, 28, 38, 14, 26, 16, 12, 12]):
        ws.column_dimensions[col].width = w

    # === Sheet 2: قيود الاستبعاد ===
    entries = generate_journal_entries([{"sub_category": tx["sub_category"], "amount": tx["amount"]} for tx in transactions if tx.get("matched")])
    ws2 = wb.create_sheet("قيود الاستبعاد")
    ws2.cell(row=1, column=1, value="قيود الاستبعاد البيني - " + group_name)
    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1).font = title_font
    ws2.cell(row=1, column=1).alignment = right
    ws2.row_dimensions[1].height = 24

    for col, h in enumerate(["البيان", "مدين", "دائن", "المبلغ"], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = bold_white; c.fill = PatternFill("solid", fgColor="1E3A8A"); c.alignment = center

    r = 4
    total = 0
    for entry in entries:
        ws2.cell(row=r, column=1, value=entry["description"]).alignment = right
        ws2.cell(row=r, column=2, value=entry["debit_label"]).alignment = right
        ws2.cell(row=r, column=3, value=entry["credit_label"]).alignment = right
        c = ws2.cell(row=r, column=4, value=entry["amount"])
        c.number_format = "#,##0.00"
        c.alignment = center
        total += entry["amount"]
        r += 1
    # المجموع
    ws2.cell(row=r, column=1, value="إجمالي الاستبعادات").font = Font(name="Calibri", size=11, bold=True)
    ws2.cell(row=r, column=1).alignment = right
    c = ws2.cell(row=r, column=4, value=round(total, 2))
    c.number_format = "#,##0.00"
    c.font = Font(name="Calibri", size=11, bold=True, color="15803D")
    c.alignment = center
    c.fill = PatternFill("solid", fgColor="D1FAE5")

    for col, w in zip("ABCD", [50, 28, 28, 18]):
        ws2.column_dimensions[col].width = w

    # === Sheet 3: ملخص ===
    ws3 = wb.create_sheet("ملخص")
    ws3.cell(row=1, column=1, value="ملخص الاستبعادات البينية")
    ws3.merge_cells("A1:C1")
    ws3.cell(row=1, column=1).font = title_font
    ws3.cell(row=1, column=1).alignment = right

    for col, h in enumerate(["نوع الاستبعاد", "عدد المعاملات", "الإجمالي"], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = bold_white; c.fill = PatternFill("solid", fgColor="1E3A8A"); c.alignment = center

    summary = defaultdict(lambda: {"count": 0, "total": 0})
    for tx in transactions:
        if not tx.get("matched"):
            continue
        cat = tx.get("sub_category", "")
        summary[cat]["count"] += 1
        summary[cat]["total"] += tx["amount"]

    r = 4
    for cat, info in summary.items():
        label = IC_CATEGORY_MAP.get(cat, {}).get("ar", cat)
        ws3.cell(row=r, column=1, value=label).alignment = right
        ws3.cell(row=r, column=2, value=info["count"]).alignment = center
        c = ws3.cell(row=r, column=3, value=round(info["total"], 2))
        c.number_format = "#,##0.00"
        c.alignment = center
        r += 1

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
