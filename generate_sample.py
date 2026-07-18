"""
Generate a balanced sample trial-balance Excel file for testing.
Numbers are tuned so:
  - Total Dr = Total Cr (balanced TB)
  - Assets = Liabilities + Equity + Net Profit (balanced BS)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

# All numbers hand-tuned so:
#   Σ Dr = Σ Cr
#   Assets (net) = Liab + Equity (incl. Net Profit)
#
# Asset side (debit):
#   Current:    Cash 80k, Bank 220k, AR 150k, Notes 25k, Inv 75k, Prep 10k = 560k
#   Non-current: Furniture 35k, Computers 28k, Vehicles 90k, Software 15k, Inv 60k = 228k
#   Sub-total: 788k
#   Less accum dep:    28k
#   Net assets:        760k
#
# Liability + Equity:
#   Liab:  Payables 90k + Notes 35k + ST loans 60k + Bank od 40k + End-of-svc 22k
#          + Accruals 15k + LT loans 120k = 382k
#   Equity: Capital 200k + Reserves 37k + RE (opening 50k) = 287k
#   Sub-total: 669k
#   Net profit (Rev - Exp): 760 - 669 = 91k
#
# Revenue:  Sales 750k + Services 95k + Interest 4k = 849k
# Expenses: COGS 360k + Selling 34k + Admin 165.5k + Depr 13k + Finance 14k + Other 4k = 590.5k
# Profit:   849 - 590.5 = 258.5k   ← wait that's more than 91k target
#
# Let me recompute target profit and adjust expenses accordingly.

assets_net = 760_000
liab = 382_000
capital = 200_000
reserves = 37_000
re_opening = 50_000
equity_contrib = capital + reserves + re_opening
profit_target = assets_net - liab - equity_contrib
print(f"Target net profit = {profit_target:,}")

# Revenue:
revenue = 750_000 + 95_000 + 4_000
expense_needed = revenue - profit_target
print(f"Revenue = {revenue:,}, Expense needed = {expense_needed:,}")

# Build expense lines (sum to expense_needed)
expense_lines = [
    ("5101", "تكلفة المبيعات",                     360_000),
    ("5201", "مصاريف دعاية وإعلان",                20_000),
    ("5202", "عمولات مبيعات",                      14_000),
    ("5301", "الرواتب والأجور",                   120_000),
    ("5302", "إيجار المكتب",                       30_000),
    ("5303", "كهرباء وماء",                         7_000),
    ("5304", "هاتف وبريد",                          3_500),
    ("5305", "صيانة",                               5_000),
    ("5401", "إهلاك الأثاث",                        1_000),
    ("5402", "إهلاك الأجهزة",                       3_000),
    ("5403", "إهلاك السيارات",                      9_000),
    ("5501", "فوائد مدينة",                        12_000),
    ("5502", "رسوم بنكية",                          2_000),
    ("5901", "مصروفات متنوعة",                      4_000),
]
sum_exp = sum(v for _, _, v in expense_lines)
print(f"Sum of expense lines: {sum_exp:,} (need {expense_needed:,}, diff {sum_exp - expense_needed:,})")

# If off, scale the largest line to match
diff = sum_exp - expense_needed
if diff != 0:
    # Adjust "الرواتب والأجور" to absorb the diff
    for i, (c, n, v) in enumerate(expense_lines):
        if c == "5301":
            expense_lines[i] = (c, n, v - diff)
            break

sum_exp = sum(v for _, _, v in expense_lines)
print(f"Adjusted expenses: {sum_exp:,}")
assert sum_exp == expense_needed, f"Expense mismatch: {sum_exp} != {expense_needed}"


# ── Build the full TB
template = [
    # ASSETS — current
    ("1101", "الصندوق",                                80_000, 0),
    ("1102", "البنك الأهلي - حساب جاري",             220_000, 0),
    ("1201", "المدينون",                              150_000, 0),
    ("1202", "أوراق القبض",                            25_000, 0),
    ("1301", "المخزون",                                75_000, 0),
    ("1401", "مصروفات مقدمة",                          10_000, 0),
    # ASSETS — non-current
    ("1601", "الأثاث",                                 35_000, 0),
    ("1602", "أجهزة الحاسب الآلي",                     28_000, 0),
    ("1603", "السيارات",                               90_000, 0),
    ("1691", "مجمع إهلاك الأثاث",                           0,  4_000),
    ("1692", "مجمع إهلاك الأجهزة",                          0,  6_000),
    ("1693", "مجمع إهلاك السيارات",                         0, 18_000),
    ("1701", "برامج الحاسب",                           15_000, 0),
    ("1801", "استثمارات في شركات",                     60_000, 0),
    # LIABILITIES
    ("2101", "الدائنون",                                    0,  90_000),
    ("2102", "أوراق الدفع",                                 0,  35_000),
    ("2201", "قرض قصير الأجل",                             0,  60_000),
    ("2202", "سلفة بنكية",                                  0,  40_000),
    ("2301", "مخصص تعويضات نهاية الخدمة",                   0,  22_000),
    ("2401", "مصروفات مستحقة",                              0,  15_000),
    ("2501", "قرض طويل الأجل",                              0, 120_000),
    # EQUITY
    ("3101", "رأس المال",                                   0, 200_000),
    ("3201", "الاحتياطي النظامي",                           0,  25_000),
    ("3202", "احتياطي عام",                                 0,  12_000),
    ("3301", "أرباح مبقاة (رصيد أول المدة)",                0, re_opening),
    # REVENUE
    ("4101", "إيرادات المبيعات",                            0, 750_000),
    ("4201", "إيرادات الخدمات",                             0,  95_000),
    ("4901", "إيرادات فوائد",                               0,   4_000),
]
# Add expenses
for code, name, val in expense_lines:
    template.append((code, name, val, 0))


# ── Verify balance
total_dr = sum(d for _, _, d, _ in template)
total_cr = sum(c for _, _, _, c in template)
print(f"\nTotal Dr = {total_dr:,}, Total Cr = {total_cr:,}, Diff = {total_dr - total_cr:,}")
assert total_dr == total_cr, "TB not balanced"

assets_debit = sum(d for c, _, d, _ in template if c.startswith(("11", "12", "13", "14", "16", "17", "18")))
contra_credit = sum(cr for c, _, _, cr in template if c.startswith("169"))
net_assets = assets_debit - contra_credit

liab_cr = sum(cr for c, _, _, cr in template if c.startswith(("21", "22", "23", "24", "25")))
equity_cr = sum(cr for c, _, _, cr in template if c.startswith(("31", "32", "33")))
rev_cr = sum(cr for c, _, _, cr in template if c.startswith(("41", "42", "49")))
exp_dr = sum(d for c, _, d, _ in template if c.startswith("5"))
net_profit = rev_cr - exp_dr
total_eq_with_profit = equity_cr + net_profit
print(f"Net assets = {net_assets:,}")
print(f"Liab       = {liab_cr:,}")
print(f"Equity + NP= {total_eq_with_profit:,}")
print(f"Balanced?  {net_assets == liab_cr + total_eq_with_profit}")
assert net_assets == liab_cr + total_eq_with_profit, "BS doesn't balance"


# ── Write Excel
out_dir = Path(__file__).parent / "samples"
out_dir.mkdir(exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "ميزان المراجعة"

ws["A1"] = "شركة المثال التجارية"
ws["A1"].font = Font(name="Cairo", size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", readingOrder=2)
ws.merge_cells("A1:D1")

ws["A2"] = "ميزان المراجعة كما في 31 ديسمبر 2024"
ws["A2"].font = Font(name="Cairo", size=12, italic=True)
ws["A2"].alignment = Alignment(horizontal="center", readingOrder=2)
ws.merge_cells("A2:D2")

headers = ["رمز الحساب", "اسم الحساب", "مدين", "دائن"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(name="Cairo", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E293B")
    c.alignment = Alignment(horizontal="center", readingOrder=2)

row_num = 5
for code, name, dr, cr in template:
    ws.cell(row=row_num, column=1, value=code)
    ws.cell(row=row_num, column=2, value=name)
    if dr != 0:
        ws.cell(row=row_num, column=3, value=dr)
    if cr != 0:
        ws.cell(row=row_num, column=4, value=cr)
    for col in range(1, 5):
        c = ws.cell(row=row_num, column=col)
        c.font = Font(name="Cairo", size=11)
        c.alignment = Alignment(horizontal="center" if col in (1, 3, 4) else "right", readingOrder=2)
        c.number_format = '#,##0.00;[Red](#,##0.00);"—"' if col in (3, 4) else 'General'
    row_num += 1

# Totals
ws.cell(row=row_num, column=1, value="الإجمالي")
ws.cell(row=row_num, column=1).font = Font(name="Cairo", size=12, bold=True, color="FFFFFF")
ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor="0F172A")
ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right", readingOrder=2)
ws.cell(row=row_num, column=2, value="")
ws.cell(row=row_num, column=2).fill = PatternFill("solid", fgColor="0F172A")
ws.cell(row=row_num, column=3, value=total_dr)
ws.cell(row=row_num, column=3).font = Font(name="Cairo", size=12, bold=True, color="FFFFFF")
ws.cell(row=row_num, column=3).fill = PatternFill("solid", fgColor="0F172A")
ws.cell(row=row_num, column=3).number_format = '#,##0.00'
ws.cell(row=row_num, column=4, value=total_cr)
ws.cell(row=row_num, column=4).font = Font(name="Cairo", size=12, bold=True, color="FFFFFF")
ws.cell(row=row_num, column=4).fill = PatternFill("solid", fgColor="0F172A")
ws.cell(row=row_num, column=4).number_format = '#,##0.00'

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

out = out_dir / "sample_trial_balance.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
